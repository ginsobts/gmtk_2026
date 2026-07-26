#!/usr/bin/env python3
"""
把游戏当前的 .txt 配置表反向生成一个可编辑的 Excel（game_tables.xlsx）。
用来「初始化」策划用的 Excel：跑一次就得到一个填好现有内容、列也对好的表，
之后只在 Excel 里改，再用 export_tables.py（或 Unity 菜单）导回 txt 即可。

用法：
    pip install openpyxl
    python tools/import_tables.py                 # 默认输出 tools/game_tables.xlsx

说明：
- 读取 Assets/Resources/GameData/<sheet>.txt，跳过空行与 # 注释行，首个数据行视为表头。
- 单元格里的字面量 \\n 会还原成真正的换行，方便在 Excel 里编辑多行文本。
- 若目标 xlsx 已存在，为避免覆盖你的改动，会改写到 game_tables.imported.xlsx 并提示。
"""
import os
import sys

SHEETS = ("characters", "dialogue", "rounds", "ui", "phases", "npc_dialogues", "portraits", "spawns", "phase_spawns", "choices")
SRC_DIR = os.path.join("Assets", "Resources", "GameData")


def read_txt(path):
    """返回 (rows)：每行是按 Tab 切好的单元格列表；跳过空行与 # 注释。"""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.rstrip("\n").rstrip("\r")
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            cells = [c.replace("\\n", "\n") for c in line.split("\t")]
            rows.append(cells)
    return rows


def main(out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("需要 openpyxl：请先运行  pip install openpyxl")
        sys.exit(1)

    if os.path.exists(out_path):
        alt = os.path.join(os.path.dirname(out_path), "game_tables.imported.xlsx")
        print(f"目标已存在：{out_path}\n为避免覆盖你的改动，改写到：{alt}")
        out_path = alt

    wb = Workbook()
    wb.remove(wb.active)

    made = 0
    for name in SHEETS:
        src = os.path.join(SRC_DIR, name + ".txt")
        if not os.path.exists(src):
            print(f"跳过：找不到 {src}")
            continue
        rows = read_txt(src)
        if not rows:
            print(f"跳过：{src} 无有效数据")
            continue

        ws = wb.create_sheet(title=name)
        for r, cells in enumerate(rows, start=1):
            for c, val in enumerate(cells, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
                if "\n" in str(val):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        # 简单加宽列，便于阅读
        for c in range(1, len(rows[0]) + 1):
            ws.column_dimensions[chr(64 + c) if c <= 26 else "A"].width = 26
        print(f"生成 sheet {name}: {len(rows)} 行（含表头）")
        made += 1

    if made == 0:
        print("没有生成任何 sheet，请检查 GameData/*.txt 是否存在。")
        sys.exit(1)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"完成：{out_path}\n在 Excel 里编辑后，用 export_tables.py（或 Unity 菜单 GMTK/配置表：Excel → txt）导回。")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join("tools", "game_tables.xlsx")
    main(out)
