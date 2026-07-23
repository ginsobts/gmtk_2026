#!/usr/bin/env python3
"""
把策划维护的 Excel 表导出为游戏读取的 Tab 分隔 .txt 表。

用法：
    pip install openpyxl
    python tools/export_tables.py                       # 默认读 tools/game_tables.xlsx
    python tools/export_tables.py path/to/tables.xlsx

Excel 里请建这三个工作表（sheet），首行为表头：
    characters : charId | displayName | artFolder | dialogueId
    dialogue   : dialogueId | order | line
    rounds     : roundId | npcCount | imposterCount | film | assign

说明：
- 输出到 Assets/Resources/GameData/<sheet>.txt，Tab 分隔，UTF-8。
- 单元格里的真实换行会被写成字面量 \\n（游戏读取时还原）。
- 只导出上面三个已知 sheet；其它 sheet 忽略。
"""
import os
import sys

SHEETS = ("characters", "dialogue", "rounds")
OUT_DIR = os.path.join("Assets", "Resources", "GameData")


def cell_to_text(value):
    if value is None:
        return ""
    s = str(value)
    # 去掉 Excel 数值型的 .0 尾巴（order/count 常见）
    if isinstance(value, float) and value.is_integer():
        s = str(int(value))
    # 真实换行 -> 字面量 \n；制表符 -> 空格，避免破坏列分隔
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\\n")
    s = s.replace("\t", " ")
    return s.strip()


def export(xlsx_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("需要 openpyxl：请先运行  pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(xlsx_path):
        print(f"找不到 Excel：{xlsx_path}")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)

    exported = 0
    for name in SHEETS:
        if name not in wb.sheetnames:
            print(f"跳过：Excel 里没有 sheet «{name}»")
            continue
        ws = wb[name]
        lines = []
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            cells = [cell_to_text(c) for c in row]
            # 丢掉整行为空的行
            if not any(cells):
                continue
            lines.append("\t".join(cells))
        out_path = os.path.join(OUT_DIR, name + ".txt")
        with open(out_path, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lines) + "\n")
        print(f"导出 {name}: {len(lines)} 行 -> {out_path}")
        exported += 1

    if exported == 0:
        print("没有导出任何表，请检查 sheet 名称。")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.path.join("tools", "game_tables.xlsx")
    export(xlsx)
